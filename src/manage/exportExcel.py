import openpyxl
import MySQLdb
import datetime
import numpy as np
import sys
import cv2
import os

A2Z = [chr(i) for i in range(65, 65+26)]

defaultRowHeight = 15
baseColWidth = 10

def fetchData(id):
	sql = MySQLdb.connect(
		user=os.environ['MYSQL_USER'], passwd=os.environ['MYSQL_PASSWORD'],
		host=os.environ['MYSQL_HOST'], db=os.environ['MYSQL_DATABASE'])

	cur = sql.cursor()
	cur.execute("SET NAMES utf8")

	query = "SELECT * FROM shape"
	cur.execute(query)
	shapes = cur.fetchall()

	query = "SELECT * FROM artwork WHERE `id`=" + str(id)
	cur.execute(query)
	artwork = cur.fetchone()

	query = "SELECT * FROM damage WHERE `artwork_id`=" + str(id)
	cur.execute(query)
	damage = cur.fetchall()

	query = "SELECT damage_img.id, damage_img.damage_id, damage_img.img FROM damage_img " \
		+ "JOIN damage ON damage_img.damage_id = damage.id WHERE `artwork_id` =" + str(id)
	cur.execute(query)
	damage_img = cur.fetchall()

	cur.close()
	sql.close()

	return shapes, artwork, damage, damage_img

def loadShapes(shapes):
	shape_imgs = {}
	for shape in shapes:
		shape_imgs[shape[0]] = cv2.imread('../img/shape/' + shape[2], cv2.IMREAD_UNCHANGED)
	return shape_imgs

def drawImgOnOverlay(img, overlay, x, y, add=False):
	w, h = img.shape[1], img.shape[0]
	x1 = int(x - w/2)
	y1 = int(y - h/2)
	x2 = int(x + w/2)
	y2 = int(y + h/2)
	o_x1 = 0
	o_y1 = 0
	o_x2 = w
	o_y2 = w
	if x1 < 0:
		o_x1 = -x1 + 1
		x1 = 0
	if y1 < 0:
		o_y1 = -y1 + 1
		y1 = 0
	if x2 >= overlay.shape[1]:
		o_x2 -= x2 - overlay.shape[1] + 1
		x2 = overlay.shape[1] - 1
	if y2 >= overlay.shape[0]:
		o_y2 -= y2 - overlay.shape[0] + 1
		y2 = overlay.shape[0] - 1

	if add:
		overlay[y1:y2, x1:x2] += img[o_y1:o_y2, o_x1:o_x2]
	else:
		overlay[y1:y2, x1:x2] = img[o_y1:o_y2, o_x1:o_x2]

	return overlay

def saveExcel(id):
	shapes, artwork, damages, damage_imgs = fetchData(id=id)
	shape_imgs = loadShapes(shapes)

	wb = openpyxl.Workbook()
	ws = wb.worksheets[0]

	ws['A2'] = '美術品名'
	ws['B2'] = artwork[1]
	ws['A3'] = 'タグ'
	ws['B3'] = artwork[2]
	ws['A3'] = 'コメント'
	ws['B3'] = artwork[3]
	ws['A4'] = '最終更新'
	ws['B4'] = artwork[5]
	ws.column_dimensions['B'].width = 20

	length = 100

	img = cv2.imread('../img/artwork/' + artwork[4])
	overlay = np.zeros((img.shape[0], img.shape[1], 4), dtype=np.uint8)

	marker_size = int(20 * max(img.shape[1], img.shape[0]) / 1000)
	text_size = int(max(img.shape[1], img.shape[0]) / 1000)

	# 損傷を描画
	for damage in damages:
		x, y = damage[8], damage[9]
		radius = damage[10]
		if radius > 0:
			color = damage[6][1:]
			color = np.array(tuple(int(color[i:i+2], 16) for i in (4, 2, 0))) / 255

			shape_img = shape_imgs[damage[7]]
			draw_img = np.zeros((shape_img.shape[0], shape_img.shape[1], 3), np.uint8)
			for i in range(draw_img.shape[2]):
				draw_img[:,:,i] = (shape_img[:,:,3] * color[i]).astype(np.uint8)
			draw_img = np.dstack((draw_img, shape_img[:,:,3:] // 2))
			draw_img = cv2.resize(draw_img, (int(radius * 2), int(radius * 2)))

			drawImgOnOverlay(draw_img, overlay, x, y, True)

	for damage in damages:
		x, y = damage[8], damage[9]

		color = damage[6][1:]
		color = np.array(tuple(int(color[i:i+2], 16) for i in (4, 2, 0))) / 255

		shape_img = shape_imgs[damage[7]]
		draw_img = np.zeros((shape_img.shape[0], shape_img.shape[1], 3), np.uint8)
		margin = int(shape_img.shape[0] * 0.1)
		for i in range(draw_img.shape[2]):
			resize_img = cv2.resize(shape_img[:,:,3], (shape_img.shape[1]-margin*2, shape_img.shape[0]-margin*2))
			draw_img[margin:-margin,margin:-margin,i] = (resize_img * color[i]).astype(np.uint8)
		draw_img = np.dstack((draw_img, shape_img[:,:,3:]))
		draw_img = cv2.resize(draw_img, (marker_size, marker_size))

		drawImgOnOverlay(draw_img, overlay, x, y)

	for damage in damages:
		x, y = damage[8], damage[9]
		w, h = text_size * 10, text_size * 10
		cv2.putText(overlay, str(damage[0]), (int(x+w/2+text_size), int(y-h/2+text_size)), cv2.FONT_HERSHEY_PLAIN, int(text_size), (0, 0, 0, 255), text_size, cv2.LINE_AA)
		cv2.putText(overlay, str(damage[0]), (int(x+w/2), int(y-h/2)), cv2.FONT_HERSHEY_PLAIN, int(text_size), (255, 255, 255, 255), text_size, cv2.LINE_AA)

	scale = baseColWidth * length / img.shape[1]
	img = cv2.resize(img, None, fx=scale, fy=scale)
	overlay = cv2.resize(overlay, None, fx=scale, fy=scale)

	cv2.imwrite('tmp/artwork_img.png', img)
	cv2.imwrite('tmp/overlay.png', overlay)

	img = openpyxl.drawing.image.Image('tmp/artwork_img.png')
	ws.add_image(img, 'D2')
	img = openpyxl.drawing.image.Image('tmp/overlay.png')
	ws.add_image(img, 'D2')

	# 参考画像を追加
	wb.create_sheet()
	ws = wb.worksheets[1]
	ws.column_dimensions['B'].width = 20

	row_count = 1
	row_step = 10
	height = 10
	for damage in damages:
		col_count = 0

		ws['A' + str(row_count+1)] = 'ID'
		ws['B' + str(row_count+1)] = damage[0]
		ws['A' + str(row_count+2)] = damage[2]
		ws['A' + str(row_count+3)] = damage[3]
		ws['A' + str(row_count+4)] = '登録日'
		ws['B' + str(row_count+4)] = damage[4]
		ws['A' + str(row_count+5)] = '削除日'
		ws['B' + str(row_count+5)] = damage[5]

		for damage_img in damage_imgs:
			if damage[0] == damage_img[1]:
				img = cv2.imread('../img/damage/' + damage_img[2])
				if img is None:
					continue
				scale = defaultRowHeight * height / img.shape[0]
				img = cv2.resize(img, None, fx=scale, fy=scale)
				cv2.imwrite('tmp/' + damage_img[2], img)
				img2 = openpyxl.drawing.image.Image('tmp/' + damage_img[2])
				ws.add_image(img2, A2Z[col_count+3] + str(row_count+1))
				col_count += int(img.shape[1] / baseColWidth / defaultRowHeight) * 3

		row_count += row_step

	dt_now = datetime.datetime.now()

	artwork_name = artwork[1].replace(' ', '_').replace('/', '').replace('\\', '')
	fname = 'tmp/export_%d_%s_%s.xlsx' % (artwork[0], artwork_name, dt_now.strftime('%Y-%m-%d %H-%M-%S'))
	wb.save(fname)
	return fname

if __name__ == '__main__':
	for i in range(1, len(sys.argv)):
		fname = saveExcel(int(sys.argv[i]))
		print(fname)

